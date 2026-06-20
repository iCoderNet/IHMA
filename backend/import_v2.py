"""
Bulk import script for docs/v2 Excel files.

Usage (from backend/ directory):
    python import_v2.py
    python import_v2.py --year 2025 --month 1

All 13 data sheets across 6 Excel files are imported.
Idempotent: run twice → creates sections/columns once, skips existing rows only
if you pass --clear to wipe rows before re-importing.
"""
import asyncio
import argparse
import os
import sys
import io
from pathlib import Path

# ── Make app importable ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))

# Load .env before importing app modules
try:
    from dotenv import load_dotenv
    load_dotenv(BASE_DIR / ".env")
except ImportError:
    pass  # python-dotenv optional if env vars already set

import openpyxl
from sqlalchemy import select, delete, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import engine, AsyncSessionLocal, Base
from app.core.database import _run_migrations
from app.models.section import Section, SectionColumn, SectionRow, SectionCell, Bolim
from app.models.district import District
# Import all models so metadata is complete
from app.models import ijtimoiy_hodim as _ih  # noqa
from app.models import user as _u  # noqa


# ═══════════════════════════════════════════════════════════════════════════════
# District Cyrillic → Latin mapping
# ═══════════════════════════════════════════════════════════════════════════════
DOCS_DIR = BASE_DIR.parent.parent / "docs" / "v2"

DISTRICT_LATIN = {
    'андижон тумани':    'Andijon tumani',
    'андижон шаҳри':     'Andijon shahri',
    'асака тумани':      'Asaka',
    'балиқчи тумани':    'Baliqchi',
    'булоқбоши тумани':  'Buloqboshi',
    'бўстон тумани':     "Bo'ston",
    'жалақудуқ тумани':  'Jalolquduq',
    'жалолқудуқ тумани': 'Jalolquduq',
    'избоскан тумани':   'Izboskan',
    'мархамат тумани':   'Marhamat',
    'олтинкўл тумани':   "Oltinko'l",
    'пахтаобод тумани':  'Paxtaobod',
    'улуғнор тумани':    "Ulug'nor",
    'хонобод шаҳри':     'Xonobod shahri',
    'хўжаобод тумани':   "Xo'jaobod",
    'шахрихон тумани':   'Shahrixon',
    'қўрғонтепа тумани': "Qo'rg'ontepa",
}

# Districts missing from initial seed
EXTRA_DISTRICTS = ['Andijon tumani', 'Buloqboshi', 'Xonobod shahri']


# ═══════════════════════════════════════════════════════════════════════════════
# Section definitions
# Each entry: (file, sheet, name, full_name, icon, color, bolim_name, cols)
# cols: list of (excel_col_index, display_name, key, data_type)
# ═══════════════════════════════════════════════════════════════════════════════
BOLIM_DEFS = [
    {"name": "НБШФБ",        "full_name": "Nogironligi bo'lgan shaxslar va fuqarolar bilan ishlash bo'limi", "icon": "♿", "color": "#3B82F6", "order": 1},
    {"name": "Болалар",       "full_name": "Bolalar shu'basi",                                                "icon": "👶", "color": "#10B981", "order": 2},
    {"name": "Тўловлар",      "full_name": "Ijtimoiy to'lovlar",                                              "icon": "💰", "color": "#F59E0B", "order": 3},
    {"name": "Мониторинг",    "full_name": "Ijtimoiy xizmatlarni monitoring qilish",                          "icon": "📊", "color": "#8B5CF6", "order": 4},
    {"name": "Инклюзив",      "full_name": "Inklyuziv ta'lim",                                                "icon": "📚", "color": "#EF4444", "order": 5},
    {"name": "Катта ёш",      "full_name": "Katta yosh bo'limi",                                              "icon": "👴", "color": "#06B6D4", "order": 6},
]

SECTIONS_CONFIG = [
    # ── НБШФБ.xlsx ─────────────────────────────────────────────────────────────
    {
        "file": "НБШФБ.xlsx",
        "sheet": "НБШ ",
        "name": "НБШ",
        "full_name": "Nogironligi bo'lgan shaxslar",
        "icon": "♿", "color": "#3B82F6", "bolim": "НБШФБ", "order": 1,
        "cols": [
            (3, "Jami НБШ",  "jami",    "number"),
            (4, "НББ",       "nbb",     "number"),
            (5, "I guruh",   "guruh_1", "number"),
            (6, "II guruh",  "guruh_2", "number"),
            (7, "III guruh", "guruh_3", "number"),
        ],
    },
    {
        "file": "НБШФБ.xlsx",
        "sheet": "РТВ ва ПОМ",
        "name": "РТВ ва ПОМ",
        "full_name": "RTV va POM mahsulotlari bilan ta'minlanganlar",
        "icon": "🦽", "color": "#6366F1", "bolim": "НБШФБ", "order": 2,
        "cols": [
            (3, "Jami",  "jami", "number"),
            (4, "РТВ",   "rtv",  "number"),
            (5, "ПОМ",   "pom",  "number"),
        ],
    },
    {
        "file": "НБШФБ.xlsx",
        "sheet": "Бандлик",
        "name": "Бандлик",
        "full_name": "Bandligi ta'minlangan NBSh",
        "icon": "💼", "color": "#0EA5E9", "bolim": "НБШФБ", "order": 3,
        "cols": [
            (3, "Jami",    "jami",    "number"),
            (4, "I guruh", "guruh_1", "number"),
            (5, "II guruh","guruh_2", "number"),
            (6, "III guruh","guruh_3","number"),
        ],
    },
    {
        "file": "НБШФБ.xlsx",
        "sheet": "Реабилитацияга борганлар",
        "name": "Реабилитация",
        "full_name": "Reabilitatsiyaga borganlar",
        "icon": "🏥", "color": "#14B8A6", "bolim": "НБШФБ", "order": 4,
        "cols": [
            (3, "Jami",    "jami",    "number"),
            (4, "НББ",     "nbb",     "number"),
            (5, "I guruh", "guruh_1", "number"),
            (6, "II guruh","guruh_2", "number"),
            (7, "III guruh","guruh_3","number"),
        ],
    },
    # ── Болалар шуъбаси.xlsx ───────────────────────────────────────────────────
    {
        "file": "Болалар шуъбаси.xlsx",
        "sheet": "Васийлик ва хомийлик (Болалар)",
        "name": "Васийлик",
        "full_name": "Vasiylik va homiylik",
        "icon": "🤝", "color": "#10B981", "bolim": "Болалар", "order": 5,
        "cols": [
            (3, "Jami",      "jami",     "number"),
            (4, "Vasiylik",  "vasiylik", "number"),
            (5, "Homiylik",  "homiylik", "number"),
            (6, "Patronat",  "patronat", "number"),
        ],
    },
    {
        "file": "Болалар шуъбаси.xlsx",
        "sheet": "Етим болалар",
        "name": "Етим болалар",
        "full_name": "Etim va ota-ona qaramoqsiz bolalar",
        "icon": "🧒", "color": "#059669", "bolim": "Болалар", "order": 6,
        "cols": [
            (3, "Jami",           "jami",          "number"),
            (4, "Etim bolalar",   "etim",          "number"),
            (5, "Ota-ona mahrum", "ota_ona_mahrum","number"),
        ],
    },
    # ── Ижтимоий тўловлар.xlsx ────────────────────────────────────────────────
    {
        "file": "Ижтимоий тўловлар.xlsx",
        "sheet": "Ижтимоий реестр (Тўловлар)",
        "name": "Ижтимоий реестр",
        "full_name": "Ijtimoiy reestr",
        "icon": "📋", "color": "#F59E0B", "bolim": "Тўловлар", "order": 7,
        "cols": [
            (3, "Jami",            "jami",           "number"),
            (4, "Davlat ta'minot", "davlat_taminot", "number"),
            (5, "Kambag'al",       "kambagal",       "number"),
            (6, "Kambag'allik chegara", "chegara",   "number"),
            # cols 7,8 are lookup helpers → skip
        ],
    },
    {
        "file": "Ижтимоий тўловлар.xlsx",
        "sheet": "Нафақа (Тўловлар)",
        "name": "Нафақа",
        "full_name": "Nafaqa oluvchilar",
        "icon": "💵", "color": "#D97706", "bolim": "Тўловлар", "order": 8,
        "cols": [
            (3, "Jami",           "jami",           "number"),
            (4, "Bolalar nafaqasi","bolalar_nafaqa", "number"),
            (5, "Moddiy yordam",  "moddiy_yordam",  "number"),
            # cols 6-9 are lookup helpers → skip
        ],
    },
    # ── Ижтимоий_хизматларни_мониторинг_қилиш.xlsx ───────────────────────────
    {
        "file": "Ижтимоий_хизматларни_мониторинг_қилиш.xlsx",
        "sheet": "Саховат ва кўмак",
        "name": "Саховат ва кўмак",
        "full_name": "Saxovat va ko'mak jamg'armasi",
        "icon": "🎁", "color": "#8B5CF6", "bolim": "Мониторинг", "order": 9,
        "cols": [
            (3, "Jami",       "jami",       "number"),
            (4, "Kiyim-bosh", "kiyim",      "number"),
            (5, "Oziq-ovqat", "oziq_ovqat", "number"),
            (6, "Dori-darmon","dori",       "number"),
            (7, "Uy-joy",     "uy_joy",     "number"),
            (8, "Tibbiy",     "tibbiy",     "number"),
        ],
    },
    # ── Инклюзив таълим.xlsx ──────────────────────────────────────────────────
    {
        "file": "Инклюзив таълим.xlsx",
        "sheet": "Таълим билан қамраб олинганлар",
        "name": "Таълим",
        "full_name": "Ta'lim bilan qamrab olinganlar",
        "icon": "📚", "color": "#EF4444", "bolim": "Инклюзив", "order": 10,
        "cols": [
            (3,  "Umumiy",              "umumiy",              "number"),
            (4,  "Qamrab olinganlar",   "qarab_olinganlar",    "number"),
            (5,  "Maktab",              "maktab",              "number"),
            (6,  "Uy ta'lim",           "uy_talim",            "number"),
            (7,  "Maxsus maktab",       "maxsus_maktab",       "number"),
            (8,  "Maxsus maktabgacha",  "maxsus_maktabgacha",  "number"),
            (9,  "Kunduzgi",            "kunduzgi",            "number"),
            (10, "Qamrab olinmaganlar", "qamrab_olinmaganlar", "number"),
        ],
    },
    # ── Катта ёш.xlsx ─────────────────────────────────────────────────────────
    {
        "file": "Катта ёш.xlsx",
        "sheet": "Тазйиқ (Катта ёш ва Болалар)",
        "name": "Тазйиқ",
        "full_name": "Tazyiq va zo'ravonlikka uchragan shaxslar",
        "icon": "🛡️", "color": "#DC2626", "bolim": "Катта ёш", "order": 11,
        "cols": [
            (3, "Jami",         "jami",         "number"),
            (4, "Ayollar",      "ayollar",      "number"),
            (5, "Bolalar",      "bolalar",      "number"),
            (6, "Himoya order", "himoya_order", "number"),
        ],
    },
    {
        "file": "Катта ёш.xlsx",
        "sheet": "Фаол хаётга қадам (Катта ёш)",
        "name": "Фаол хаётга қадам",
        "full_name": "Faol hayotga qadam dasturi",
        "icon": "🚶", "color": "#06B6D4", "bolim": "Катта ёш", "order": 12,
        "cols": [
            (3, "Jami",                   "jami",                  "number"),
            (4, "Ijtimoiy-maishiy",       "ijtimoiy_maishiy",      "number"),
            (5, "Qarab turish (kunduzgi)","qarab_turish_kunduzi",  "number"),
            (6, "Tibbiy reabilitatsiya",  "tibbiy_reabilitatsiya", "number"),
            (7, "Qarab turish",           "qarab_turish",          "number"),
            (8, "Hamrohlik",              "hamrohlik",             "number"),
        ],
    },
    {
        "file": "Катта ёш.xlsx",
        "sheet": "САНАТОРИЯ (Катта ёш)",
        "name": "Санатория",
        "full_name": "Agentlik tizimidagi sanatoriya",
        "icon": "🏨", "color": "#0891B2", "bolim": "Катта ёш", "order": 13,
        "cols": [
            (3, "Jami",           "jami",       "number"),
            (4, "Yoshga doir",    "yoshga_doir","number"),
            (5, "I-II guruh",     "guruh_1_2",  "number"),
            (6, "Urush qatnashchilari", "urush","number"),
        ],
    },
]


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def safe_num(val) -> str | None:
    """Convert Excel cell to string number or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', '-', '—'):
        return None
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return None


def is_skip_row(row: list) -> bool:
    """True for empty rows, Жами totals, or header artifacts."""
    if not any(row):
        return True
    v0 = str(row[0]).strip() if row[0] is not None else ''
    v1 = str(row[1]).strip() if row[1] is not None else ''
    v2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
    if v1 in ('Жами', 'жами', '#'):
        return True
    if v0 == '#':
        return True
    if not v0.isdigit():
        return True
    # Skip rows where BOTH district AND MFY name are empty (ghost/padding rows)
    if not v1 and not v2:
        return True
    return False


def read_data_rows(fpath: str, sheet_name: str) -> list:
    """Read data rows: skip first 2 empty rows, skip header rows (R2,R3), skip Жами (R4)."""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    # Strip trailing space from sheet names
    ws = None
    for sn in wb.sheetnames:
        if sn.strip() == sheet_name.strip():
            ws = wb[sn]
            break
    if ws is None:
        wb.close()
        raise ValueError(f"Sheet not found: '{sheet_name}' in {fpath}")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # rows 0,1 = empty; row 2 = main header; row 3 = sub-header; row 4 = Жами total
    # data starts at row 5
    data = []
    for row in all_rows[5:]:
        row_list = list(row)
        if is_skip_row(row_list):
            continue
        data.append(row_list)
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# Main import logic
# ═══════════════════════════════════════════════════════════════════════════════

async def ensure_districts(db: AsyncSession) -> dict[str, int]:
    """Return {latin_name_lower: district_id}, creating missing ones."""
    result = await db.execute(select(District))
    existing = {d.name: d.id for d in result.scalars().all()}

    for name in EXTRA_DISTRICTS:
        if name not in existing:
            d = District(name=name, name_ru=name)
            db.add(d)
            print(f"  ➕ Creating district: {name}")
    await db.commit()

    result = await db.execute(select(District))
    return {d.name.lower(): d.id for d in result.scalars().all()}


def resolve_district(cyrillic_name: str, latin_map: dict[str, int]) -> int | None:
    """Map Cyrillic or Latin district name → district_id (3-stage lookup)."""
    key = cyrillic_name.strip().lower()
    if not key:
        return None
    # Stage 1: direct Latin match (in case input is already Latin)
    if key in latin_map:
        return latin_map[key]
    # Stage 2: Cyrillic → Latin via DISTRICT_LATIN
    latin = DISTRICT_LATIN.get(key)
    if latin and latin.lower() in latin_map:
        return latin_map[latin.lower()]
    # Stage 3: partial match (substring)
    for db_name, db_id in latin_map.items():
        if db_name in key or key in db_name:
            return db_id
    return None


async def ensure_bolimlar(db: AsyncSession) -> dict[str, int]:
    """Create Bo'lim records if missing. Return {name: id}."""
    result = await db.execute(select(Bolim))
    existing = {b.name: b.id for b in result.scalars().all()}

    for bdef in BOLIM_DEFS:
        if bdef["name"] not in existing:
            b = Bolim(**bdef)
            db.add(b)
            print(f"  ➕ Creating bo'lim: {bdef['name']}")
    await db.commit()

    result = await db.execute(select(Bolim))
    return {b.name: b.id for b in result.scalars().all()}


async def ensure_section(
    db: AsyncSession,
    cfg: dict,
    bolim_map: dict[str, int],
    order_offset: int,
) -> tuple[int, dict[str, int]]:
    """Get or create section + columns. Return (section_id, {col_key: col_id})."""
    result = await db.execute(select(Section).where(Section.name == cfg["name"]))
    section = result.scalar_one_or_none()

    bolim_id = bolim_map.get(cfg["bolim"])

    if section is None:
        section = Section(
            name=cfg["name"],
            full_name=cfg["full_name"],
            icon=cfg["icon"],
            color=cfg["color"],
            bolim_id=bolim_id,
            order=cfg["order"],
        )
        db.add(section)
        await db.commit()
        await db.refresh(section)
        print(f"  ➕ Created section: {cfg['name']}")
    else:
        # Update bolim_id if missing
        if section.bolim_id is None and bolim_id:
            section.bolim_id = bolim_id
            await db.commit()

    # Ensure columns
    col_result = await db.execute(
        select(SectionColumn).where(SectionColumn.section_id == section.id)
    )
    existing_cols = {c.key: c.id for c in col_result.scalars().all()}

    for order_i, (_, col_name, col_key, col_type) in enumerate(cfg["cols"]):
        if col_key not in existing_cols:
            col = SectionColumn(
                section_id=section.id,
                name=col_name,
                key=col_key,
                data_type=col_type,
                order=order_i,
            )
            db.add(col)
    await db.commit()

    col_result = await db.execute(
        select(SectionColumn).where(SectionColumn.section_id == section.id)
    )
    col_map = {c.key: c.id for c in col_result.scalars().all()}
    return section.id, col_map


async def import_section(
    db: AsyncSession,
    cfg: dict,
    section_id: int,
    col_map: dict[str, int],
    district_map: dict[str, int],
    period_year: int | None,
    period_month: int | None,
    clear_existing: bool,
):
    """Import all data rows for one section."""
    fpath = str(DOCS_DIR / cfg["file"])
    if not os.path.exists(fpath):
        print(f"  ⚠️  File not found: {fpath}")
        return 0

    if clear_existing:
        # Collect row IDs first (FK constraint: delete cells before rows)
        row_q = select(SectionRow.id).where(SectionRow.section_id == section_id)
        if period_year is not None:
            row_q = row_q.where(SectionRow.period_year == period_year)
        if period_month is not None:
            row_q = row_q.where(SectionRow.period_month == period_month)
        row_ids = (await db.execute(row_q)).scalars().all()
        if row_ids:
            await db.execute(delete(SectionCell).where(SectionCell.row_id.in_(row_ids)))
            await db.execute(delete(SectionRow).where(SectionRow.id.in_(row_ids)))
        await db.commit()
        period_str = f"{period_year}/{period_month}" if period_year else "barcha"
        print(f"  🗑  Cleared {len(row_ids)} rows for {cfg['name']} (period: {period_str})")

    data_rows = read_data_rows(fpath, cfg["sheet"])

    # Count existing rows to generate order
    count_res = await db.execute(
        select(SectionRow).where(SectionRow.section_id == section_id)
    )
    row_offset = len(count_res.scalars().all())

    rows_added = 0
    batch_rows = []
    batch_cells = []

    for ri, excel_row in enumerate(data_rows):
        # Resolve district from col[1]
        district_cyrillic = str(excel_row[1]).strip() if excel_row[1] else ''
        mfy_name = str(excel_row[2]).strip() if len(excel_row) > 2 and excel_row[2] else ''

        district_id = resolve_district(district_cyrillic, district_map)

        row = SectionRow(
            section_id=section_id,
            district_id=district_id,
            mfy_name=mfy_name,
            order=row_offset + ri,
            period_year=period_year,
            period_month=period_month,
        )
        db.add(row)
        await db.flush()  # get row.id

        # MFY stored as text cell — find the "mfy" column if present
        # (We don't add mfy as a defined column since it's used as district label)
        # We'll just store it in a special column if one exists
        # Actually: just store the number columns defined in cfg["cols"]

        for (excel_col_idx, _, col_key, _) in cfg["cols"]:
            col_id = col_map.get(col_key)
            if col_id is None:
                continue
            val = excel_row[excel_col_idx] if excel_col_idx < len(excel_row) else None
            num_str = safe_num(val)
            cell = SectionCell(row_id=row.id, column_id=col_id, value=num_str)
            db.add(cell)

        rows_added += 1

        # Commit in batches of 200
        if rows_added % 200 == 0:
            await db.commit()
            print(f"    … {rows_added}/{len(data_rows)} rows")

    await db.commit()
    return rows_added


async def main(year: int | None, month: int | None, clear: bool):
    print("\n🚀 Starting v2 bulk import")
    print(f"   Period: year={year}, month={month}")
    print(f"   Clear existing: {clear}")
    print(f"   Source: {DOCS_DIR}\n")

    # Ensure schema exists
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await _run_migrations(conn)
    print("✅ Schema ready\n")

    async with AsyncSessionLocal() as db:
        print("📍 Ensuring districts…")
        district_map = await ensure_districts(db)
        print(f"   {len(district_map)} districts available\n")

        print("📁 Ensuring bo'limlar…")
        bolim_map = await ensure_bolimlar(db)
        print(f"   {len(bolim_map)} bo'limlar available\n")

        total_imported = 0

        for cfg in SECTIONS_CONFIG:
            print(f"📊 Section: {cfg['name']} ← {cfg['file']} / {cfg['sheet'].strip()}")
            section_id, col_map = await ensure_section(db, cfg, bolim_map, cfg["order"])

            count = await import_section(
                db, cfg, section_id, col_map,
                district_map, year, month, clear,
            )
            print(f"   ✅ {count} rows imported\n")
            total_imported += count

    print(f"🎉 Done! Total rows imported: {total_imported}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bulk import docs/v2 Excel data")
    parser.add_argument("--year",  type=int, default=None, help="Period year (e.g. 2025)")
    parser.add_argument("--month", type=int, default=None, choices=range(1, 13), help="Period month 1-12")
    parser.add_argument("--clear", action="store_true", help="Delete existing rows before re-importing")
    args = parser.parse_args()

    asyncio.run(main(args.year, args.month, args.clear))
