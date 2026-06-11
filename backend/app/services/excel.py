"""Excel import/export service for dynamic sections."""
import io
import re
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_CYRILLIC_MAP = str.maketrans({
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo',
    'ж':'j','з':'z','и':'i','й':'y','к':'k','л':'l','м':'m',
    'н':'n','о':'o','п':'p','р':'r','с':'s','т':'t','у':'u',
    'ф':'f','х':'x','ц':'ts','ч':'ch','ш':'sh','щ':'sh',
    'ъ':'','ы':'i','ь':'','э':'e','ю':'yu','я':'ya',
    # O'zbek maxsus harflari
    'ў':'o','қ':'q','ғ':'g','ҳ':'h',
})

def slugify(text: str, max_len: int = 64) -> str:
    """Convert column name to a URL/DB-safe ASCII key (Cyrillic → Latin)."""
    text = str(text).strip().lower()
    text = text.translate(_CYRILLIC_MAP)
    text = re.sub(r'[^\w\s-]', '', text, flags=re.ASCII)
    text = re.sub(r'[\s_-]+', '_', text)
    text = text.strip('_')[:max_len].rstrip('_')
    return text or "col"


def _cell_display_value(cell) -> str | None:
    """Return human-readable cell value, converting percentages correctly."""
    if cell.value is None:
        return None
    if isinstance(cell.value, (int, float)):
        fmt = str(cell.number_format or '')
        if '%' in fmt:
            pct = cell.value * 100
            # Trim unnecessary decimals
            formatted = f"{pct:.10g}"
            return formatted + '%'
    return str(cell.value)


def _read_rows_with_format(ws) -> list[list]:
    """Read all rows preserving percentage formatting."""
    rows = []
    for row in ws.iter_rows():
        rows.append([_cell_display_value(c) for c in row])
    return rows


def read_excel_preview(file_bytes: bytes, sheet_index: int = 0, max_rows: int = 10) -> dict:
    """Read Excel file and return preview data."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames

    ws = wb.worksheets[min(sheet_index, len(wb.worksheets) - 1)]
    all_rows = _read_rows_with_format(ws)

    preview_rows = []
    for i, row in enumerate(all_rows[:max_rows + 20]):
        preview_rows.append({
            "index": i,
            "data": [c if c is not None else None for c in row]
        })

    total = len(all_rows)
    headers = [c for c in (all_rows[0] if all_rows else [])]

    wb.close()
    return {
        "total_rows": total,
        "headers": headers,
        "preview_rows": preview_rows,
        "sheet_names": sheet_names,
    }


def import_excel_rows(
    file_bytes: bytes,
    sheet_index: int,
    skip_rows: int,
    header_row_index: int,
    skip_columns: list[int],
    column_mapping: dict[int, str],  # excel_col_index -> section column key
) -> list[dict[str, Any]]:
    """
    Parse Excel and return list of row dicts {column_key: value}.
    skip_rows: rows to completely ignore at the top
    header_row_index: 0-based index (from start of file, AFTER skip) for headers
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[sheet_index]
    all_rows = _read_rows_with_format(ws)
    wb.close()

    # Apply skip
    all_rows = all_rows[skip_rows:]

    if not all_rows:
        return []

    results = []
    # Data rows = everything after header
    data_rows = all_rows[header_row_index + 1:]

    for row in data_rows:
        if all(c is None for c in row):
            continue
        row_data = {}
        for col_idx, value in enumerate(row):
            if col_idx in skip_columns:
                continue
            if col_idx in column_mapping:
                key = column_mapping[col_idx]
                row_data[key] = value if value is not None else ""
        if row_data:
            results.append(row_data)

    return results


def export_section_to_excel(
    section_name: str,
    columns: list[dict],
    rows: list[dict],
) -> bytes:
    """Generate a styled Excel file from section data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = section_name[:30]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(columns) + 1)}1")
    title_cell = ws["A1"]
    title_cell.value = section_name
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # Header row
    ws.cell(row=2, column=1, value="№").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = header_align
    ws.cell(row=2, column=1).border = border
    ws.column_dimensions["A"].width = 5

    for i, col in enumerate(columns, start=2):
        cell = ws.cell(row=2, column=i, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = max(15, len(col["name"]) + 2)

    ws.row_dimensions[2].height = 30

    # Data rows
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_i, row in enumerate(rows, start=1):
        excel_row = row_i + 2
        # Row number
        num_cell = ws.cell(row=excel_row, column=1, value=row_i)
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.border = border
        if row_i % 2 == 0:
            num_cell.fill = alt_fill

        for col_i, col in enumerate(columns, start=2):
            val = row.get("cells", {}).get(col["key"], "")
            cell = ws.cell(row=excel_row, column=col_i, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_i % 2 == 0:
                cell.fill = alt_fill

    # Freeze header
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_import_template(columns: list[dict]) -> bytes:
    """Generate an empty template Excel for importing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(col["name"]) + 4)

    ws.row_dimensions[1].height = 25

    # Add a sample row with hints
    hint_fill = PatternFill("solid", fgColor="DBEAFE")
    for i, col in enumerate(columns, start=1):
        dtype = col.get("data_type", "text")
        hint = "Matn" if dtype == "text" else ("Raqam" if dtype == "number" else "Sana")
        cell = ws.cell(row=2, column=i, value=f"[{hint}]")
        cell.fill = hint_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
