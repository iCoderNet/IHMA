"""Ijtimoiy hodimlar (Social workers) CRUD + Excel import."""
import json
import io
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.core.dependencies import get_current_user
from app.models.ijtimoiy_hodim import IjtimoiyHodim
from app.models.district import District
from app.models.user import User

router = APIRouter(prefix="/ijtimoiy-hodimlar", tags=["ijtimoiy-hodimlar"])

# Kirill → Latin tuman nomlar xaritasi
_CYR_TO_LAT: dict[str, str] = {
    'андижон тумани':    'andijon tumani',
    'андижон шаҳри':     'andijon shahri',
    'асака тумани':      'asaka',
    'балиқчи тумани':    'baliqchi',
    'булоқбоши тумани':  'buloqboshi',
    'бўстон тумани':     "bo'ston",
    'жалақудуқ тумани':  'jalolquduq',
    'жалолқудуқ тумани': 'jalolquduq',
    'избоскан тумани':   'izboskan',
    'мархамат тумани':   'marhamat',
    'олтинкўл тумани':   "oltinko'l",
    'пахтаобод тумани':  'paxtaobod',
    'улуғнор тумани':    "ulug'nor",
    'хонобод шаҳри':     'xonobod shahri',
    'хўжаобод тумани':   "xo'jaobod",
    'шахрихон тумани':   'shahrixon',
    'қўрғонтепа тумани': "qo'rg'ontepa",
}

def _resolve_district_name(raw: str, districts_lower: dict[str, int]) -> int | None:
    """Kirill yoki Latin tuman nomidan district_id topadi."""
    key = raw.strip().lower()
    # 1. To'g'ridan to'g'ri moslik
    if key in districts_lower:
        return districts_lower[key]
    # 2. Kirill → Latin o'girma orqali
    lat = _CYR_TO_LAT.get(key)
    if lat and lat in districts_lower:
        return districts_lower[lat]
    # 3. Qisman moslik (DB da qisqaroq nom bo'lishi mumkin)
    for db_name, db_id in districts_lower.items():
        if db_name in key or key in db_name:
            return db_id
    return None


# ── Schemas ───────────────────────────────────────────────────────────────────

class HodimCreate(BaseModel):
    district_id: int | None = None
    mfy_name: str
    fio: str
    phone: str = ""


class HodimUpdate(BaseModel):
    district_id: int | None = None
    mfy_name: str | None = None
    fio: str | None = None
    phone: str | None = None


class HodimOut(BaseModel):
    id: int
    district_id: int | None
    district_name: str | None = None
    mfy_name: str
    fio: str
    phone: str

    class Config:
        from_attributes = True


def _to_out(h: IjtimoiyHodim) -> dict:
    return {
        "id": h.id,
        "district_id": h.district_id,
        "district_name": h.district.name if h.district else None,
        "mfy_name": h.mfy_name,
        "fio": h.fio,
        "phone": h.phone,
    }


# ── List ──────────────────────────────────────────────────────────────────────

@router.get("")
async def list_hodimlar(
    district_id: int | None = Query(None),
    search: str | None = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = (
        select(IjtimoiyHodim)
        .options(selectinload(IjtimoiyHodim.district))
        .order_by(IjtimoiyHodim.district_id, IjtimoiyHodim.mfy_name, IjtimoiyHodim.fio)
    )
    if district_id:
        q = q.where(IjtimoiyHodim.district_id == district_id)
    if search:
        like = f"%{search}%"
        from sqlalchemy import or_
        q = q.where(or_(
            IjtimoiyHodim.fio.ilike(like),
            IjtimoiyHodim.mfy_name.ilike(like),
            IjtimoiyHodim.phone.ilike(like),
        ))

    total_q = select(func.count()).select_from(IjtimoiyHodim)
    if district_id:
        total_q = total_q.where(IjtimoiyHodim.district_id == district_id)
    total = (await db.execute(total_q)).scalar()

    q = q.offset((page - 1) * size).limit(size)
    items = (await db.execute(q)).scalars().all()

    return {
        "items": [_to_out(h) for h in items],
        "total": total,
        "page": page,
        "pages": max(1, -(-total // size)),
    }


# ── Create ────────────────────────────────────────────────────────────────────

@router.post("", status_code=201)
async def create_hodim(
    data: HodimCreate,
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    h = IjtimoiyHodim(**data.model_dump(), created_by=current.id)
    db.add(h)
    await db.commit()
    await db.refresh(h)
    # reload with district
    result = await db.execute(
        select(IjtimoiyHodim).where(IjtimoiyHodim.id == h.id)
        .options(selectinload(IjtimoiyHodim.district))
    )
    return _to_out(result.scalar_one())


# ── Update ────────────────────────────────────────────────────────────────────

@router.put("/{hodim_id}")
async def update_hodim(
    hodim_id: int,
    data: HodimUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    result = await db.execute(
        select(IjtimoiyHodim).where(IjtimoiyHodim.id == hodim_id)
        .options(selectinload(IjtimoiyHodim.district))
    )
    h = result.scalar_one_or_none()
    if not h:
        raise HTTPException(404, "Topilmadi")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(h, field, value)
    await db.commit()
    await db.refresh(h)
    result2 = await db.execute(
        select(IjtimoiyHodim).where(IjtimoiyHodim.id == hodim_id)
        .options(selectinload(IjtimoiyHodim.district))
    )
    return _to_out(result2.scalar_one())


# ── Delete ────────────────────────────────────────────────────────────────────

@router.delete("/{hodim_id}", status_code=204)
async def delete_hodim(
    hodim_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    h = (await db.execute(select(IjtimoiyHodim).where(IjtimoiyHodim.id == hodim_id))).scalar_one_or_none()
    if not h:
        raise HTTPException(404, "Topilmadi")
    await db.delete(h)
    await db.commit()


# ── Excel Import ──────────────────────────────────────────────────────────────
# Expected columns: Tuman | MFY | FIO | Tel raqami

@router.post("/excel/import")
async def import_excel(
    file: UploadFile = File(...),
    district_id: int | None = Form(None),  # optional override for all rows
    sheet_index: int = Form(0),
    skip_rows: int = Form(0),
    header_row: int = Form(0),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[min(sheet_index, len(wb.worksheets) - 1)]

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()

    all_rows = all_rows[skip_rows:]
    if not all_rows:
        return {"imported": 0}

    # Detect header row and map columns
    header = all_rows[header_row]
    col_map = {}  # key -> col_index
    for i, h in enumerate(header):
        hl = h.lower()
        if any(k in hl for k in ["туман", "tuman", "район"]):
            col_map["district"] = i
        elif any(k in hl for k in ["мфй", "mfy", "маҳалла", "mahalla"]):
            col_map["mfy"] = i
        elif any(k in hl for k in ["фио", "fio", "ф.и.о", "исм", "ism", "ходим", "xodim", "hodim"]):
            col_map["fio"] = i
        elif any(k in hl for k in ["тел", "tel", "рақам", "raqam", "phone"]):
            col_map["phone"] = i

    # Load districts for name matching
    districts_result = await db.execute(select(District))
    districts_lower = {d.name.lower(): d.id for d in districts_result.scalars().all()}

    data_rows = all_rows[header_row + 1:]
    created = 0
    for row in data_rows:
        if all(c == "" for c in row):
            continue

        fio = row[col_map["fio"]].strip() if "fio" in col_map and col_map["fio"] < len(row) else ""
        mfy = row[col_map["mfy"]].strip() if "mfy" in col_map and col_map["mfy"] < len(row) else ""
        phone = row[col_map["phone"]].strip() if "phone" in col_map and col_map["phone"] < len(row) else ""

        if not fio and not mfy:
            continue

        # Resolve district (Kirill yoki Latin nomdan)
        d_id = district_id
        if not d_id and "district" in col_map and col_map["district"] < len(row):
            raw_d = row[col_map["district"]].strip()
            d_id = _resolve_district_name(raw_d, districts_lower)

        db.add(IjtimoiyHodim(
            district_id=d_id,
            mfy_name=mfy,
            fio=fio,
            phone=phone,
            created_by=current.id,
        ))
        created += 1

    await db.commit()
    return {"imported": created, "message": f"{created} ta ijtimoiy hodim import qilindi"}


# ── Excel Export ──────────────────────────────────────────────────────────────

@router.get("/excel/export")
async def export_excel(
    district_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = (
        select(IjtimoiyHodim)
        .options(selectinload(IjtimoiyHodim.district))
        .order_by(IjtimoiyHodim.district_id, IjtimoiyHodim.mfy_name)
    )
    if district_id:
        q = q.where(IjtimoiyHodim.district_id == district_id)
    items = (await db.execute(q)).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ijtimoiy hodimlar"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1E3A5F")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["№", "Tuman", "MFY", "Ijtimoiy hodim FIO", "Tel raqami"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = border
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for ri, h in enumerate(items, 1):
        row_data = [ri, h.district.name if h.district else "", h.mfy_name, h.fio, h.phone]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri + 1, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ijtimoiy_hodimlar.xlsx"},
    )
